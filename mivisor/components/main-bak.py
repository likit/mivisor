import os
import pandas
import sqlalchemy as sa
import json
import wx, wx.adv, wx.lib
from datetime import datetime
from wx.lib.wordwrap import wordwrap
from threading import Thread
from pydispatch import dispatcher
from openpyxl import load_workbook

CLOSE_DIALOG_SIGNAL = 'close-notification-dialog'

from components.datatable import DataGrid
from components.fieldcreation import (FieldCreateDialog, OrganismFieldFormDialog,
                                      DrugRegFormDialog, IndexFieldList,
                                      HeatmapFieldList, DateRangeFieldList)

APPDATA_DIR = 'appdata'
DRUG_REGISTRY_FILE = 'drugs.json'

basepath = os.path.dirname(os.path.abspath(__file__))

drug_dict = {}
drug_df = None


def load_drug_registry():
    global drug_dict
    global drug_df
    if DRUG_REGISTRY_FILE:
        try:
            drug_df = pandas.read_json(os.path.join(APPDATA_DIR, DRUG_REGISTRY_FILE))
        except:
            return pandas.DataFrame(columns=['drug', 'abbreviation', 'group'])
        else:
            drug_dict = {}
            if drug_df.empty:
                drug_df = pandas.DataFrame(columns=['drug', 'abbreviation', 'group'])
            else:
                drug_df = drug_df.sort_values(['group'])
                for idx, row in drug_df.iterrows():
                    drug = row
                    if row['abbreviation']:
                        abbrs = [a.strip().lower() for a in row['abbreviation'].split(',')]
                    else:
                        abbrs = []
                    for ab in abbrs:
                        drug_dict[ab] = drug


class FieldAttribute():
    def __init__(self):
        self.data = {}
        self.columns = []
        self.organisms = {}

    def update_from_json(self, json_data):
        if not self.columns:
            return False

        json_data = json.loads(json_data)
        profile_cols = json_data['columns']
        profile_cols_no_agg = set([col for col in profile_cols if not
                                   col.startswith('@')])
        if profile_cols_no_agg.difference(self.columns):
            return False
        else:
            self.columns = profile_cols
            self.data = json_data['data']
            self.organisms = json_data['organisms']
            return True

    def update_from_json_for_database(self, json_data):
        """
        Update columns with data from the saved profile
        :param json_data:
        :return: Boolean
        """
        # Data must be loaded first
        if not self.columns:
            return False

        json_data = json.loads(json_data)
        profile_cols = json_data['columns']

        # columns must match
        profile_cols_no_agg = [col for col in profile_cols if not col.startswith('@')]
        assert len(set(profile_cols_no_agg).difference(set(self.columns))) == 0

        self.columns = profile_cols
        self.data = json_data['data']
        self.organisms = json_data['organisms']

        return True

    def update_from_dataframe(self, data_frame):
        self.columns = []
        for n, column in enumerate(data_frame.columns):
            self.columns.append(column)
            self.data[column] = {'name': column,
                                 'alias': column,
                                 'organism': False,
                                 'key': False,
                                 'drug': True if column.lower() in drug_dict else False,
                                 'date': False,
                                 'type': str(data_frame[column].dtype),
                                 'keep': True,
                                 'desc': "",
                                 }

    def values(self):
        return self.data.values()

    def get_column(self, colname):
        try:
            return self.data[colname]
        except KeyError as e:
            raise AttributeError(e)

    def iget_column(self, index):
        try:
            return self.columns[index]
        except IndexError:
            return None

    def get_col_index(self, colname):
        try:
            return self.columns.index(colname)
        except ValueError:
            return -1

    def is_col_aggregate(self, colname):
        if colname in self.data:
            if self.data[colname].get('aggregate'):
                return True
            else:
                return False
        else:
            raise KeyError

    def update_organisms(self, df):
        self.organisms = {}
        for idx, row in df.iterrows():
            self.organisms[row[0]] = {'genus': row[1], 'species': row[2]}


def browse(filetype='MLAB'):
    file_meta = {
        'MLAB': {
            'wildcard': "Excel files (*.xls;*xlsx)|*.xls;*.xlsx"
        },
        'CSV': {
            'wildcard': "CSV files (*.csv)|*.csv"
        }
    }
    with wx.FileDialog(None, "Open data file",
                       wildcard=file_meta[filetype]['wildcard'],
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) \
            as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return

        return fileDialog.GetPath()


def show_sheets(parent, worksheets):
    dlg = wx.SingleChoiceDialog(None,
                                "Select a worksheet", "Worksheets", worksheets)
    if dlg.ShowModal() == wx.ID_OK:
        return dlg.GetStringSelection()
    dlg.Destroy()


class NotificationBox(wx.Dialog):
    def __init__(self, parent, caption, message):
        super(NotificationBox, self).__init__(parent=parent,
                                              title=caption, size=(300, 90),
                                              style=wx.CAPTION)
        self.label = wx.StaticText(self, label=message)
        vsizer = wx.BoxSizer(wx.VERTICAL)
        vsizer.Add(self.label, 1, wx.ALL | wx.EXPAND | wx.CENTER, 20)
        self.SetSizer(vsizer)
        self.Center(wx.HORIZONTAL)

        dispatcher.connect(self.endModal, signal=CLOSE_DIALOG_SIGNAL, sender=dispatcher.Any)

    def updateLabel(self, msg):
        self.label.SetLabelText(msg)

    def endModal(self, rc):
        # Not sure why sometimes the dialog is not modal,
        # but failing to check it causes an error.
        if self.IsModal():
            self.EndModal(rc)
        else:
            return rc


class MainWindow(wx.Frame):
    def __init__(self, parent):
        super(MainWindow, self).__init__(parent)
        scr_width, scr_height = wx.DisplaySize()
        self.SetIcon(wx.Icon(os.path.join(basepath, 'icons/appicon.ico')))
        self.version_no = '2019.1.8'
        self.description = 'Faculty of Medical Technology, Mahidol University'
        self.SetTitle('Mivisor Version {}'.format(self.version_no))
        self.SetSize((int(scr_width * 0.75), int(scr_height * 0.85)))
        self.Center()

        self.current_column = None
        self.data_filepath = None
        self.profile_filepath = None
        self.db_filepath = None
        self.current_session_id = None
        self.dbengine = None
        self.data_loaded = False
        self.field_attr = FieldAttribute()
        df = pandas.DataFrame({'Name': ['Mivisor'],
                               'Version': [self.version_no],
                               'Released': ['2019-08-04'],
                               'Brought to you by': [self.description],
                               'Contact': ['likit.pre@mahidol.edu']})

        menubar = wx.MenuBar()
        fileMenu = wx.Menu()
        dataMenu = wx.Menu()
        fieldMenu = wx.Menu()
        exportMenu = wx.Menu()
        antibiogramMenu = wx.Menu()
        registryMenu = wx.Menu()
        analyzeMenu = wx.Menu()
        aboutMenu = wx.Menu()
        databaseMenu = wx.Menu()
        imp = wx.Menu()
        mlabItem = imp.Append(wx.ID_ANY, 'Excel (MLAB)')
        # csvItem = imp.Append(wx.ID_ANY, 'CSV')
        # csvItem.Enable(False)
        fileMenu.AppendSeparator()
        fileMenu.Append(wx.ID_ANY, 'I&mport', imp)
        fileMenu.AppendSeparator()

        self.loadProfileItem = fileMenu.Append(wx.ID_ANY, 'Load Profile')
        self.loadProfileItem.Enable(False)

        self.saveProfileItem = fileMenu.Append(wx.ID_ANY, 'Save Profile')
        self.saveProfileItem.Enable(False)

        exitItem = fileMenu.Append(wx.ID_EXIT, 'Quit', 'Quit Application')
        self.createFieldItem = fieldMenu.Append(wx.ID_ANY, 'Matching')
        self.createFieldItem.Enable(False)

        dataMenu.Append(wx.ID_ANY, 'New field', fieldMenu)
        self.saveToDatabaseMenuItem = dataMenu.Append(wx.ID_ANY, 'Save to database')
        self.saveToDatabaseMenuItem.Enable(False)
        self.appendToDatabaseMenuItem = dataMenu.Append(wx.ID_ANY, 'Append to database')
        self.appendToDatabaseMenuItem.Enable(False)
        dataMenu.AppendSeparator()

        self.organismItem = dataMenu.Append(wx.ID_ANY, 'Organism')
        self.organismItem.Enable(False)

        dataMenu.AppendSeparator()

        self.exportToExcelMenuItem = exportMenu.Append(wx.ID_ANY, 'To Excel')
        self.saveToFlatDbMenuItem = exportMenu.Append(wx.ID_ANY, 'Create flat database')
        self.addToFlatDbMenuItem = exportMenu.Append(wx.ID_ANY, 'Add to flat database')
        self.exportToExcelMenuItem.Enable(False)
        self.saveToFlatDbMenuItem.Enable(False)
        self.addToFlatDbMenuItem.Enable(False)
        dataMenu.Append(wx.ID_ANY, 'Export flat table', exportMenu)

        drugRegMenuItem = registryMenu.Append(wx.ID_ANY, 'Drugs')

        self.biogramDbMenuItem = antibiogramMenu.Append(wx.ID_ANY, 'Create summary report')
        self.biogramDbMenuItem.Enable(True)
        self.biogramHeatmapMenuItem = antibiogramMenu.Append(wx.ID_ANY, 'Create heatmap plot')
        self.biogramHeatmapMenuItem.Enable(True)

        analyzeMenu.Append(wx.ID_ANY, 'Antibiogram', antibiogramMenu)

        aboutMenuItem = aboutMenu.Append(wx.ID_ANY, "About the program")

        self.connectDbMenuItem = databaseMenu.Append(wx.ID_ANY, 'Connect')
        self.disconnectDbMenuItem = databaseMenu.Append(wx.ID_ANY, 'Disconnect')

        self.Bind(wx.EVT_MENU, self.onConnectDbMenuItemClick, self.connectDbMenuItem)
        self.Bind(wx.EVT_MENU, self.onDisconnectDbMenuItemClick, self.disconnectDbMenuItem)
        self.Bind(wx.EVT_MENU, lambda x: self.onSaveToDatabaseMenuItemClick(x, action='replace'),
                  self.saveToDatabaseMenuItem)
        self.Bind(wx.EVT_MENU, lambda x: self.onSaveToDatabaseMenuItemClick(x, action='append'),
                  self.appendToDatabaseMenuItem)

        menubar.Append(fileMenu, '&File')
        menubar.Append(dataMenu, '&Data')
        menubar.Append(databaseMenu, 'Database')
        menubar.Append(analyzeMenu, 'Analy&ze')
        menubar.Append(registryMenu, '&Registry')
        menubar.Append(aboutMenu, '&About')
        self.SetMenuBar(menubar)

        accel_tbl = wx.AcceleratorTable([
            (wx.ACCEL_CTRL, ord('M'), mlabItem.GetId()),
        ])
        self.SetAcceleratorTable(accel_tbl)

        import sys
        self.Bind(wx.EVT_CLOSE, lambda x: sys.exit())

        self.Bind(wx.EVT_MENU, self.on_about_menu_click, aboutMenuItem)

        self.Bind(wx.EVT_MENU, self.OnQuit, exitItem)
        self.Bind(wx.EVT_MENU, self.onLoadMLABItemClick, mlabItem)
        # self.Bind(wx.EVT_MENU, self.OnLoadCSV, csvItem)

        self.Bind(wx.EVT_MENU, self.OnCreateField, self.createFieldItem)
        self.Bind(wx.EVT_MENU, self.OnSaveProfile, self.saveProfileItem)
        self.Bind(wx.EVT_MENU, self.OnLoadProfile, self.loadProfileItem)
        self.Bind(wx.EVT_MENU, self.OnOrganismClick, self.organismItem)

        # TODO: rename OnExportRawData method
        self.Bind(wx.EVT_MENU, self.OnExportRawData, self.exportToExcelMenuItem)
        self.Bind(wx.EVT_MENU, lambda x: self.onSaveToFlatDbMenuItemClick(x, action='replace'),
                  self.saveToFlatDbMenuItem)
        self.Bind(wx.EVT_MENU, lambda x: self.onSaveToFlatDbMenuItemClick(x, action='append'),
                  self.addToFlatDbMenuItem)

        self.Bind(wx.EVT_MENU, self.on_drug_reg_menu_click, drugRegMenuItem)

        self.Bind(wx.EVT_MENU, self.onBiogramDbMenuItemClick, self.biogramDbMenuItem)
        self.Bind(wx.EVT_MENU, self.onBiogramHeatmapMenuItemClick, self.biogramHeatmapMenuItem)

        # init panels
        self.info_panel = wx.Panel(self, wx.ID_ANY)
        self.preview_panel = wx.Panel(self, wx.ID_ANY)
        self.summary_panel = wx.Panel(self, wx.ID_ANY)
        self.attribute_panel = wx.Panel(self, wx.ID_ANY)
        self.edit_panel = wx.Panel(self, wx.ID_ANY)

        # init sizers
        self.info_sizer = wx.StaticBoxSizer(wx.VERTICAL, self.info_panel, "Session Information")
        self.summary_sizer = wx.StaticBoxSizer(wx.VERTICAL, self.summary_panel, "Field Summary")
        self.field_attr_sizer = wx.StaticBoxSizer(wx.VERTICAL, self.attribute_panel, "Field Attributes")
        edit_box_sizer = wx.StaticBoxSizer(wx.HORIZONTAL, self.edit_panel, "Edit")
        self.data_grid_box_sizer = wx.StaticBoxSizer(wx.VERTICAL, self.preview_panel, "Data Preview")

        self.profile_lbl = wx.StaticText(self.info_panel, -1, "Profile filepath: {}".format(self.profile_filepath))
        self.datafile_lbl = wx.StaticText(self.info_panel, -1, "Data filepath: {}".format(self.data_filepath))
        self.dbfile_lbl = wx.StaticText(self.info_panel, -1, "Database filepath: {}".format(self.db_filepath))
        self.info_sizer.Add(self.datafile_lbl)
        self.info_sizer.Add(self.profile_lbl)
        self.info_sizer.Add(self.dbfile_lbl)

        self.data_grid = DataGrid(self.preview_panel)
        self.data_grid.set_table(df)
        self.data_grid.AutoSize()
        self.data_grid_box_sizer.Add(self.data_grid, 1, flag=wx.EXPAND | wx.ALL)

        self.key_chkbox = wx.CheckBox(self.edit_panel, -1, label="Key", name="key")
        self.drug_chkbox = wx.CheckBox(self.edit_panel, -1, label="Drug", name="drug")
        self.organism_chkbox = wx.CheckBox(self.edit_panel, -1, label="Organism", name="organism")
        self.keep_chkbox = wx.CheckBox(self.edit_panel, -1, label="Included", name="keep")
        self.date_chkbox = wx.CheckBox(self.edit_panel, -1, label="Date", name="date")
        self.field_edit_checkboxes = [self.key_chkbox, self.drug_chkbox,
                                      self.keep_chkbox, self.organism_chkbox,
                                      self.date_chkbox]
        checkbox_sizer = wx.FlexGridSizer(cols=len(self.field_edit_checkboxes), hgap=4, vgap=0)
        for chkbox in self.field_edit_checkboxes:
            checkbox_sizer.Add(chkbox)
            chkbox.Bind(wx.EVT_CHECKBOX, self.on_edit_save_button_clicked)

        checkbox_label = wx.StaticText(self.edit_panel, -1, "Marked as")
        self.field_desc = wx.TextCtrl(self.edit_panel, -1, "", style=wx.TE_MULTILINE, size=(200, 100))
        self.field_alias = wx.TextCtrl(self.edit_panel, -1, "")
        self.edit_save_button = wx.Button(self.edit_panel, -1, "Update")
        self.edit_save_button.Bind(wx.EVT_BUTTON, self.on_edit_save_button_clicked)
        alias_label = wx.StaticText(self.edit_panel, -1, "Alias")
        desc_label = wx.StaticText(self.edit_panel, -1, "Description")
        form_sizer = wx.FlexGridSizer(cols=2, hgap=2, vgap=2)
        form_sizer.AddMany([checkbox_label, checkbox_sizer])
        form_sizer.AddMany([desc_label, self.field_desc])
        form_sizer.AddMany([alias_label, self.field_alias])
        form_sizer.AddMany([wx.StaticText(self.edit_panel, -1, ""), self.edit_save_button])
        edit_box_sizer.Add(form_sizer, 1, flag=wx.ALIGN_LEFT)

        self.summary_table = wx.ListCtrl(self.summary_panel, style=wx.LC_REPORT)
        self.summary_table.InsertColumn(0, 'Field')
        self.summary_table.InsertColumn(1, 'Value')
        self.summary_sizer.Add(self.summary_table, 1, wx.EXPAND)

        self.field_attr_list = wx.ListCtrl(self.attribute_panel, style=wx.LC_REPORT)
        self.add_field_attr_list_column()
        self.field_attr_sizer.Add(self.field_attr_list, 1, wx.EXPAND)
        self.Bind(wx.EVT_LIST_ITEM_SELECTED, self.onFieldAttrListItemSelected)

        self.preview_panel.SetSizer(self.data_grid_box_sizer)
        self.attribute_panel.SetSizer(self.field_attr_sizer)
        self.summary_panel.SetSizer(self.summary_sizer)
        self.edit_panel.SetSizer(edit_box_sizer)
        self.info_panel.SetSizer(self.info_sizer)

        self.vbox = wx.BoxSizer(wx.VERTICAL)
        self.hbox = wx.BoxSizer(wx.HORIZONTAL)

        self.hbox.Add(self.edit_panel, 2, flag=wx.ALL | wx.EXPAND)
        self.hbox.Add(self.summary_panel, 1, flag=wx.ALL | wx.EXPAND)
        self.vbox.Add(self.info_panel, 0, flag=wx.EXPAND | wx.ALL)
        self.vbox.Add(self.preview_panel, 1, flag=wx.EXPAND | wx.ALL)
        self.vbox.Add(self.attribute_panel, flag=wx.ALL | wx.EXPAND)
        self.vbox.Add(self.hbox, flag=wx.ALL | wx.EXPAND | wx.ALL)
        self.SetSizer(self.vbox)

        load_drug_registry()

    def OnQuit(self, e):
        self.Close()

    def OnOrganismClick(self, event):
        columns = []
        sel_col = None
        for c in self.field_attr.columns:
            col = self.field_attr.get_column(c)
            if col['keep'] and col['organism']:
                columns.append(col['alias'])

        if not columns:
            dlg = wx.MessageDialog(None, "No organism field specified.",
                                   "Please select a field for organism.",
                                   wx.OK)
            ret = dlg.ShowModal()
            if ret == wx.ID_OK:
                return

        dlg = wx.SingleChoiceDialog(None,
                                    "Select a column", "Kept columns", columns)
        if dlg.ShowModal() == wx.ID_OK:
            sel_col = dlg.GetStringSelection()
        dlg.Destroy()
        if sel_col:
            sel_col_index = self.field_attr.get_col_index(sel_col)
            column = self.field_attr.get_column(sel_col)

            values = self.data_grid.table.df[sel_col].unique()
            fc = OrganismFieldFormDialog()
            if not self.field_attr.organisms:
                _df = pandas.DataFrame({column['alias']: values, 'genus': None, 'species': None})
            else:
                orgs = []
                genuses = []
                species = []
                for org in self.field_attr.organisms:
                    orgs.append(org)
                    genuses.append(self.field_attr.organisms[org]['genus'])
                    species.append(self.field_attr.organisms[org]['species'])

                _df = pandas.DataFrame({column['alias']: orgs, 'genus': genuses, 'species': species})

            fc.grid.set_table(_df)
            resp = fc.ShowModal()
            self.field_attr.update_organisms(fc.grid.table.df)

    def load_profile_from_filepath(self, df):
        try:
            fp = open(self.profile_filepath, 'r')
        except IOError:
            wx.MessageDialog(self,
                             'Cannot read data from {}. Please double check the file path.'.format(
                                 self.profile_filepath),
                             'The profile file cannot be loaded',
                             wx.ICON_ERROR).ShowModal()
            return

        json_data = fp.read()
        fp.close()
        if not self.field_attr.update_from_json_for_database(json_data):
            wx.MessageDialog(self,
                             'Fields in the profile and the data do not match.',
                             'The profile cannot be loaded',
                             wx.ICON_INFORMATION).ShowModal()
            return

        for c in self.field_attr.columns:
            if self.field_attr.is_col_aggregate(c):
                column = self.field_attr.get_column(c)
                column_index = self.field_attr.get_col_index(c)
                if c not in df.columns:
                    d = []
                    from_col = column['aggregate']['from']
                    dict_ = column['aggregate']['data']
                    for value in df[from_col]:
                        d.append(dict_.get(value, value))
                    df.insert(column_index, c, value=d)

        return df

    def OnLoadProfile(self, event):
        if not self.data_filepath:
            dlg = wx.MessageDialog(None,
                                   "No data for this session.",
                                   "Please provide data for this session first.",
                                   wx.OK | wx.CENTER)
            ret = dlg.ShowModal()
            return

        wildcard = "JSON (*.json)|*.json"
        _profile_pth = self.data_filepath or os.getcwd()
        with wx.FileDialog(None, "Choose a file", _profile_pth,
                           "", wildcard, wx.FC_OPEN) as file_dlg:
            if file_dlg.ShowModal() == wx.ID_CANCEL:
                return
            try:
                fp = open(file_dlg.GetPath(), 'r')
                json_data = fp.read()
                fp.close()
                if not self.field_attr.update_from_json(json_data):
                    wx.MessageDialog(self,
                                     'Fields in the profile and the data do not match.',
                                     'The profile cannot be loaded',
                                     wx.ICON_INFORMATION).ShowModal()
                    return

                for c in self.field_attr.columns:
                    if self.field_attr.is_col_aggregate(c):
                        column = self.field_attr.get_column(c)
                        column_index = self.field_attr.get_col_index(c)
                        if c not in self.data_grid.table.df.columns:
                            d = []
                            from_col = column['aggregate']['from']
                            dict_ = column['aggregate']['data']
                            for value in self.data_grid.table.df[from_col]:
                                d.append(dict_.get(value, value))
                            self.data_grid.table.df.insert(column_index, c, value=d)
                            self.data_grid.table.InsertCols(column_index, 1)

                self.refresh_field_attr_list_column()
                self.update_edit_panel(self.field_attr.iget_column(0))
                self.profile_filepath = file_dlg.GetPath()
                self.profile_lbl.SetLabelText("Profile filepath: {}".format(self.profile_filepath))
            except IOError:
                print('Cannot load data from file.')

    def OnSaveProfile(self, event):
        wildcard = "JSON (*.json)|*.json"
        with wx.FileDialog(None, "Choose a file to save a profile.", os.getcwd(),
                           "", wildcard, wx.FC_SAVE) as file_dlg:
            if file_dlg.ShowModal() == wx.ID_CANCEL:
                return
            try:
                fp = open(file_dlg.GetPath(), 'w')
                for col in self.field_attr.columns:
                    column = self.field_attr.get_column(col)
                fp.write(json.dumps({'data': self.field_attr.data,
                                     'columns': self.field_attr.columns,
                                     'organisms': self.field_attr.organisms},
                                    indent=2))
                fp.close()
                self.profile_filepath = file_dlg.GetPath()
                self.profile_lbl.SetLabelText("Profile filepath: {}".format(self.profile_filepath))
            except IOError:
                print('Cannot save data to file.')

    def load_datafile(self, filetype='MLAB'):
        filepath = browse(filetype)
        if filepath and filepath:
            try:
                worksheets = load_workbook(filepath).sheetnames
            except FileNotFoundError:
                wx.MessageDialog(self,
                                 'Cannot download the data file.\nPlease check the file path again.',
                                 'File Not Found!', wx.OK | wx.CENTER).ShowModal()
            else:
                if len(worksheets) > 1:
                    sel_worksheet = show_sheets(self, worksheets)
                else:
                    sel_worksheet = worksheets[0]

                bag = {'data': None, 'filepath': ''}

                def read_excel():
                    # TODO: need to handle an error
                    df = pandas.read_excel(filepath, sheet_name=sel_worksheet)
                    bag['data'] = df
                    bag['filepath'] = filepath
                    wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=0)

                thread = Thread(target=read_excel)
                thread.start()
                with NotificationBox(self, caption='Import Data',
                                     message='Reading from the Excel file...') as md:
                    md.ShowModal()

                return bag['data'], bag['filepath']

        else:
            wx.MessageDialog(self, 'File path is not valid!',
                             'Please check the file path.',
                             wx.OK | wx.CENTER).ShowModal()
            return pandas.DataFrame(), ''

    def onLoadMLABItemClick(self, e):
        if self.data_loaded:
            dlg = wx.MessageDialog(None, "Click \"Yes\" to continue or click \"No\" to return to your session.",
                                   "Data in this current session will be discarded!",
                                   wx.YES_NO | wx.ICON_QUESTION)
            ret_ = dlg.ShowModal()
            if ret_ == wx.ID_NO:
                return
            self.profile_filepath = None
            self.db_filepath = None
            self.dbengine = None
            self.dbfile_lbl.SetLabelText('Database filepath:')
            self.profile_lbl.SetLabelText('Profile filepath:')

        df, filepath = self.load_datafile()
        if filepath:
            if df.empty:
                dlg = wx.MessageDialog(None,
                                       "Do you want to proceed?\nClick \"Yes\" to continue or \"No\" to cancel.",
                                       "Warning: dataset is empty.",
                                       wx.YES_NO | wx.ICON_QUESTION)
                ret_ = dlg.ShowModal()
                if ret_ == wx.ID_NO:
                    return

            self.data_filepath = filepath
            self.datafile_lbl.SetLabelText("Data filepath: {}".format(self.data_filepath))
            self.data_loaded = True
            self.data_grid_box_sizer.Remove(0)
            self.data_grid.Destroy()
            self.data_grid = DataGrid(self.preview_panel)
            self.data_grid.set_table(df)
            self.data_grid.AutoSizeColumns()
            self.data_grid_box_sizer.Add(self.data_grid, 1, flag=wx.EXPAND | wx.ALL)
            self.data_grid_box_sizer.Layout()  # repaint the sizer
            self.field_attr.update_from_dataframe(df)
            self.field_attr_list.ClearAll()
            self.refresh_field_attr_list_column()
            if self.field_attr.columns:
                self.current_column = self.field_attr.iget_column(0)
                self.field_attr_list.Select(0)
            self.saveProfileItem.Enable(True)
            self.loadProfileItem.Enable(True)
            self.organismItem.Enable(True)
            self.loadProfileItem.Enable(True)
            self.createFieldItem.Enable(True)
            self.appendToDatabaseMenuItem.Enable(True)
            self.saveToDatabaseMenuItem.Enable(True)
            # need to enable load profile menu item here
            # after refactoring the menu bar

    def OnLoadCSV(self, e):
        filepath = browse('CSV')
        if filepath:
            try:
                df = pandas.read_csv(filepath)
            except FileNotFoundError:
                wx.MessageDialog(self, 'Cannot download the data file.\nPlease check the file path again.',
                                 'File Not Found!', wx.OK | wx.CENTER).ShowModal()
        else:
            wx.MessageDialog(self, 'No File Path Found!',
                             'Please enter/select the file path.',
                             wx.OK | wx.CENTER).ShowModal()

    def OnCreateField(self, event):
        columns = []
        for c in self.field_attr.columns:
            col = self.field_attr.get_column(c)
            if col['keep']:
                columns.append(col['alias'])

        with wx.SingleChoiceDialog(
                None, "Select a column", "Kept columns", columns) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                sel_col = dlg.GetStringSelection()
            else:
                return

        if sel_col:
            sel_col_index = self.field_attr.get_col_index(sel_col)

            values = self.data_grid.table.df[sel_col].unique()
            _df = pandas.DataFrame({'Value': values, 'Group': values})
            with FieldCreateDialog() as fc:
                fc.grid.set_table(_df)
                resp = fc.ShowModal()

                if resp == wx.ID_OK:
                    _agg_dict = {}
                    for idx, row in fc.grid.table.df.iterrows():
                        _agg_dict[row['Value']] = row['Group']

                    _agg_data = []
                    for value in self.data_grid.table.df[sel_col]:
                        _agg_data.append(_agg_dict[value])
                    new_col = '@' + fc.field_name.GetValue()
                    if new_col in self.field_attr.columns:
                        new_col += '-copy'
                    self.data_grid.table.df.insert(sel_col_index + 1, new_col, value=_agg_data)
                    self.data_grid.AutoSize()
                    self.data_grid_box_sizer.Layout()
                    self.data_grid.table.InsertCols(sel_col_index + 1, 1)
                    self.field_attr.columns.insert(sel_col_index + 1, new_col)
                    self.field_attr.data[new_col] = {
                        'name': new_col,
                        'alias': new_col,
                        'organism': False,
                        'key': False,
                        'drug': False,
                        'date': False,
                        'type': str(self.data_grid.table.df[new_col].dtype),
                        'keep': True,
                        'desc': "",
                        'aggregate': {
                            'from': sel_col,
                            'data': _agg_dict
                        }
                    }
                    self.refresh_field_attr_list_column()
            # self.OnSaveProfile(None)
            # self.onSaveToDatabaseMenuItemClick(None)

    def reset_summary_table(self, desc):
        self.summary_table.ClearAll()
        self.summary_table.InsertColumn(0, 'Field')
        self.summary_table.InsertColumn(1, 'Value')
        for n, k in enumerate(desc.keys()):
            self.summary_table.InsertItem(n, k)
            self.summary_table.SetItem(n, 1, str(desc[k]))

    def update_edit_panel(self, colname):
        for cb in self.field_edit_checkboxes:
            name = cb.GetName()
            cb.SetValue(self.field_attr.get_column(colname)[name])

        self.field_alias.SetValue(self.field_attr.get_column(colname)['alias'])
        self.field_desc.SetValue(self.field_attr.get_column(colname)['desc'])
        self.current_column = colname

    def onFieldAttrListItemSelected(self, evt):
        index = evt.GetIndex()
        current_column = self.data_grid.table.df.columns[index]
        desc = self.data_grid.table.df[self.current_column].describe()
        self.reset_summary_table(desc=desc)
        self.update_edit_panel(current_column)
        self.data_grid.SelectCol(index)

    def refresh_field_attr_list_column(self):
        self.add_field_attr_list_column()
        self.update_field_attrs()

    def add_field_attr_list_column(self):
        self.field_attr_list.ClearAll()
        self.field_attr_list.InsertColumn(0, 'Field name')
        self.field_attr_list.InsertColumn(1, 'Alias name')
        self.field_attr_list.InsertColumn(2, 'Type')
        self.field_attr_list.InsertColumn(3, 'Key')
        self.field_attr_list.InsertColumn(4, 'Date')
        self.field_attr_list.InsertColumn(5, 'Organism')
        self.field_attr_list.InsertColumn(6, 'Drug')
        self.field_attr_list.InsertColumn(7, 'Description')
        self.field_attr_list.InsertColumn(8, 'Kept')
        self.field_attr_list.SetColumnWidth(7, 300)

    def update_field_attrs(self):
        for n, c in enumerate(self.field_attr.columns):
            col = self.field_attr.get_column(c)
            self.field_attr_list.InsertItem(n, col['name'])
            self.field_attr_list.SetItem(n, 1, col['alias'])
            self.field_attr_list.SetItem(n, 2, col['type'])
            self.field_attr_list.SetItem(n, 3, str(col['key']))
            self.field_attr_list.SetItem(n, 4, str(col['date']))
            self.field_attr_list.SetItem(n, 5, str(col['organism']))
            self.field_attr_list.SetItem(n, 6, str(col['drug']))
            self.field_attr_list.SetItem(n, 7, str(col['desc']))
            self.field_attr_list.SetItem(n, 8, str(col['keep']))

    def on_edit_save_button_clicked(self, event):
        col_index = self.field_attr.get_col_index(self.current_column)
        for cb in self.field_edit_checkboxes:
            name = cb.GetName()
            self.field_attr.get_column(self.current_column)[name] = cb.GetValue()
        self.field_attr.get_column(self.current_column)['alias'] = self.field_alias.GetValue()
        self.field_attr.get_column(self.current_column)['desc'] = self.field_desc.GetValue()
        self.refresh_field_attr_list_column()
        self.field_attr_list.Select(col_index)
        self.field_attr_list.Focus(col_index)

    def convert_to_flat(self, engine, startdate, enddate, deduplicate=True):
        info_columns = []
        dup_keys = []
        organism_column = None
        for colname in self.field_attr.columns:
            column = self.field_attr.get_column(colname)
            if column['keep']:
                if column['key'] and not column['organism'] and not column['drug'] and deduplicate:
                    dup_keys.append(colname)
                if column['organism']:
                    organism_column = column
                elif column['date']:
                    date_column = colname
                    info_columns.append(column)
                elif column['drug']:
                    continue
                else:
                    info_columns.append(column)

        if not organism_column:
            with wx.MessageDialog(self,
                                  "Please specify the organism column.",
                                  "Export failed.",
                                  wx.OK) as md:
                md.ShowModal()
            wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=1)

        rf = pandas.read_sql_table('records', con=engine)
        df = pandas.read_sql_table('drugs', con=engine)

        dict_ = {}
        for column in info_columns:
            dict_[column['alias']] = self.data_grid.table.df[column['name']]

        dict_['sur_key'] = rf['sur_key']

        genuses = []
        species = []
        organisms = []
        for org in self.data_grid.table.df[organism_column['name']]:
            organisms.append(org)
            org_item = self.field_attr.organisms.get(org, {'genus': org, 'species': org})
            genuses.append(org_item.get('genus', org))
            species.append(org_item.get('species', org))

        dict_[organism_column['alias']] = organisms
        dict_['genus'] = genuses
        dict_['species'] = species
        dict_['organism_name'] = [' '.join(item) for item in zip(genuses, species)]

        def get_drug_group(x):
            global drug_dict
            return drug_dict.get(x.lower(), pandas.Series()).get('group', 'unspecified')

        exported_data = pandas.DataFrame(dict_)

        if deduplicate:
            if not dup_keys:
                with wx.MessageDialog(self,
                                      "Please specify some key columns.",
                                      "Export failed.",
                                      wx.OK) as md:
                    md.ShowModal()
                wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=1)
            else:
                #TODO: inform user about error in deduplication if no date was found..
                dup_keys.append('organism_name')
                # dup_keys.append('drug')
                if dup_keys and date_column:
                    exported_data = exported_data.sort_values(by=date_column)
                    exported_data = exported_data.drop_duplicates(
                        subset=dup_keys, keep='first'
                    )
                else:
                    with wx.MessageDialog(self,
                                          "Please specify a date column.",
                                          "Export failed.",
                                          wx.OK) as md:
                        md.ShowModal()
                    wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=1)

        df['drugGroup'] = df['drug'].apply(lambda x: get_drug_group(x))
        self.flat_dataframe = exported_data.merge(df, on='sur_key', how='inner')
        del self.flat_dataframe['sur_key']  # remove surrogate key column

        if startdate and enddate:
            try:
                self.flat_dataframe = self.flat_dataframe[
                    (self.flat_dataframe[date_column] >= startdate) & (self.flat_dataframe[date_column] <= enddate)]
            except TypeError:
                wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=1)

        wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=0)


    def OnExportRawData(self, event):
        wildcard = "Excel (*.xlsx;*.xls)|*.xlsx;*.xls"
        with wx.FileDialog(None, "Choose a file", os.getcwd(),
                           "", wildcard, wx.FC_SAVE) as file_dlg:
            if file_dlg.ShowModal() == wx.ID_CANCEL:
                return
            else:
                output_filepath = file_dlg.GetPath()

        date_dlg = DateRangeFieldList(self)

        if date_dlg.ShowModal() == wx.ID_OK:
            deduplicate = date_dlg.deduplicate.IsChecked()
            if not date_dlg.all.IsChecked():
                startdate = map(int, date_dlg.startDatePicker.GetValue().FormatISODate().split('-'))
                enddate = map(int, date_dlg.endDatePicker.GetValue().FormatISODate().split('-'))
                startdate = pandas.Timestamp(*startdate)
                enddate = pandas.Timestamp(*enddate)
            else:
                startdate = None
                enddate = None

        thread = Thread(target=self.convert_to_flat, args=(self.dbengine, startdate, enddate, deduplicate))
        thread.start()
        with NotificationBox(self, caption='Export Data',
                             message='Preparing data to export...') as nd:
            result = nd.ShowModal()

        if result > 0:
            return

        '''
        for colname in self.field_attr.columns:
            column = self.field_attr.get_column(colname)
            if column['keep'] and column['date']:
                date_column = colname
        '''

        # df = self.flat_dataframe
        def write_to_excel(flat_df, output_filepath):
            try:
                flat_df.to_excel(output_filepath, engine='xlsxwriter', index=False)
            except:
                with wx.MessageDialog(None,
                                        "Cannot save data to the output file.",
                                        "Export failed.",
                                        wx.OK) as md:
                    md.ShowModal()
                wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=1)
            else:
                with wx.MessageDialog(None,
                                        "Data have been export to Excel as a flat table.",
                                        "Export succeeds.",
                                        wx.OK) as md:
                    md.ShowModal()
                wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=0)

        thread = Thread(target=write_to_excel, args=(self.flat_dataframe, output_filepath))
        thread.start()
        with NotificationBox(self, caption='Writing Data',
                             message='Writing data to Excel file...') as nd:
            result = nd.ShowModal()

        if result > 0:
            return


    def onSaveToFlatDbMenuItemClick(self, event, action='replace'):
        style = wx.FD_SAVE
        if not self.profile_filepath:
            wx.MessageDialog(None, "No profile path specified.",
                             "Please save a profile to a file or load a profile to the session before continue.",
                             wx.OK).ShowModal()
            return

        # Select date range to export data
        date_dlg = DateRangeFieldList(self)

        if date_dlg.ShowModal() == wx.ID_OK:
            deduplicate = date_dlg.deduplicate.IsChecked()
            if not date_dlg.all.IsChecked():
                startdate = map(int, date_dlg.startDatePicker.GetValue().FormatISODate().split('-'))
                enddate = map(int, date_dlg.endDatePicker.GetValue().FormatISODate().split('-'))
                startdate = pandas.Timestamp(*startdate)
                enddate = pandas.Timestamp(*enddate)
            else:
                startdate = None
                enddate = None

        if self.dbengine:
            thread = Thread(target=self.convert_to_flat, args=(self.dbengine, startdate, enddate, deduplicate))
            thread.start()
            with NotificationBox(self, caption='Export Data',
                                message='Preparing data to export...'
                                ) as nd:
                result = nd.ShowModal()
            if result == 1:
                wx.MessageDialog(None, "Could not save data to the database.",
                                "Export failed.",
                                wx.OK).ShowModal()
            if result == 2:
                wx.MessageDialog(None, "Could not save the profile data to the database.",
                                "Export failed.",
                                wx.OK).ShowModal()
        else:
            wx.MessageDialog(None, "Please first save data to the database.",
                            "Export failed.",
                            wx.OK).ShowModal()
            return

        with wx.FileDialog(None, "Choose an SQLite data file",
                            wildcard='SQLite files (*.sqlite;*.db)|*.sqlite;*.db',
                            style=style) \
                as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            else:
                dw_filepath = fileDialog.GetPath()
        if dw_filepath:
            dwengine = sa.create_engine('sqlite:///{}'.format(dw_filepath))

        try:
            self.flat_dataframe['added_at'] = datetime.utcnow()
            self.flat_dataframe.to_sql('facts', con=dwengine, if_exists=action, index=False)
        except IOError:
            wx.MessageDialog(self, "Error occurred while saving the data to the database.",
                                "Failed to save the data.",
                                wx.OK).ShowModal()
            return

        metadata = pandas.DataFrame({'profile': [self.profile_filepath], 'updatedAt': [datetime.utcnow()]})

        try:
            metadata.to_sql('metadata', con=dwengine, if_exists='replace', index=False)
        except IOError:
            wx.MessageDialog(self, "Error occurred while saving the metadata to the database.",
                                "Failed to save the metadata.",
                                wx.OK).ShowModal()
            return

        wx.MessageDialog(self, "Data have been exported to the database.",
                            "Finished.",
                            wx.OK).ShowModal()

    def onSaveToDatabaseMenuItemClick(self, event, action='replace'):
        if not self.profile_filepath:
            with wx.MessageDialog(None, message='Please save the profile to a file first.',
                                  caption='Profile file not found error.',
                                  style=wx.OK | wx.CENTER) as msgDialog:
                msgDialog.ShowModal()
            return

        if action == 'append' or not self.dbengine:
            with wx.FileDialog(None, "Choose or specify a database file",
                               wildcard='SQLite files (*.sqlite;*.db)|*.sqlite;*.db',
                               style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) \
                    as fileDialog:
                if fileDialog.ShowModal() == wx.ID_CANCEL:
                    return
                else:
                    self.db_filepath = fileDialog.GetPath()

        if self.db_filepath:
            if action == 'replace':
                with wx.MessageDialog(None,
                                      "Are you sure you want to write to {}".format(self.db_filepath),
                                      "Database is about to be overwritten.",
                                      wx.OK | wx.CANCEL) as msgDialog:
                    ret = msgDialog.ShowModal()
                    if ret == wx.ID_CANCEL:
                        return

            elif action == 'append':
                with wx.MessageDialog(None,
                                      "Are you sure you want to write to {}".format(self.db_filepath),
                                      "Database is about to be modified.",
                                      wx.OK | wx.CANCEL) as msgDialog:
                    ret = msgDialog.ShowModal()
                    if ret == wx.ID_CANCEL:
                        return

            metadata = pandas.DataFrame({'profile': [self.profile_filepath], 'updatedAt': [datetime.utcnow()]})
            self.dbfile_lbl.SetLabelText('Database filepath {} CONNECTED'.format(self.db_filepath))
            self.dbengine = sa.create_engine('sqlite:///{}'.format(self.db_filepath))

            # add surrogate keys
            try:
                records_df = pandas.read_sql_table('records', con=self.dbengine)
            except ValueError:
                sur_key_start = 0
            else:
                sur_key_start = len(records_df)

            # generate surrogate keys based on existing records
            if action == 'append':
                db_metadata = pandas.read_sql_table('metadata', con=self.dbengine)
                db_profile_filepath = db_metadata.tail(1)['profile'].tolist()[0]
                if db_profile_filepath != self.profile_filepath:
                    with wx.MessageDialog(None,
                                          "Cannot tell whether the data structure is compatible."
                                          "\nPlease use the same profile.".format(self.db_filepath),
                                          "Check database schema.",
                                          wx.OK | wx.CANCEL) as msgDialog:
                        ret = msgDialog.ShowModal()
                        return

                sur_keys = range(sur_key_start, sur_key_start + len(self.data_grid.table.df))
            else:
                sur_keys = range(0, len(self.data_grid.table.df))

            # TODO: use .loc[row_indexer,col_indexer] = value instead
            self.data_grid.table.df['sur_key'] = sur_keys

            # split data into records and drugs
            rec_columns = [c for c in self.field_attr.columns
                           if self.field_attr.data[c]['drug'] is False
                           and not self.field_attr.is_col_aggregate(c)]
            rec_columns += ['sur_key']
            drug_columns = [c for c in self.field_attr.columns
                            if self.field_attr.data[c]['drug'] is True] + ['sur_key']
            records_frame = self.data_grid.table.df[rec_columns]
            drugs_frame = self.data_grid.table.df[drug_columns]
            drugs_frame.fillna('-', inplace=True)
            drugs_frame = drugs_frame.set_index('sur_key')\
                .stack().reset_index().rename(columns={'level_1': 'drug', 0: 'sensitivity'})
            # save records into records table
            # stack drug data using surrogate keys and reset the indexes then rename columns and save to drugs table
            try:
                records_frame.to_sql('records', con=self.dbengine, index=False, if_exists=action)
                drugs_frame.to_sql('drugs', con=self.dbengine, index=False, if_exists=action)
                metadata.to_sql('metadata', con=self.dbengine, if_exists='replace', index=False)
            except:
                with wx.MessageDialog(None, message='Failed to save data to the database.',
                                      caption='Data saving failed.',
                                      style=wx.OK | wx.CENTER) as msgDialog:
                    msgDialog.ShowModal()
            else:
                with wx.MessageDialog(None, message='Data have been saved to the database file.', caption='Finished.',
                                      style=wx.OK | wx.CENTER) as msgDialog:
                    msgDialog.ShowModal()
                self.exportToExcelMenuItem.Enable(True)
                self.saveToFlatDbMenuItem.Enable(True)
                self.addToFlatDbMenuItem.Enable(True)

    def on_drug_reg_menu_click(self, event):
        global drug_df
        # TODO: drug table should be sortable by all columns
        drug_filepath = os.path.join(APPDATA_DIR, DRUG_REGISTRY_FILE)
        dr = DrugRegFormDialog()
        dr.grid.set_table(drug_df)
        dr.grid.AutoSize()
        resp = dr.ShowModal()
        # TODO: values not saved until the cell is unfocused
        if resp == wx.ID_OK:
            dr.grid.table.df.to_json(drug_filepath)
            drug_df = dr.grid.table.df.copy()

    def on_about_menu_click(self, event):
        info = wx.adv.AboutDialogInfo()
        info.Name = "Mivisor"
        info.Version = self.version_no
        info.Copyright = "(C) 2019 Faculty of Medical Technology, Mahidol University"
        info.Description = wordwrap(self.description + "\n" +
                                    "For more information, please go to http://mtclan.net/mivisor",
                                    500, wx.ClientDC(self.preview_panel))
        info.WebSite = ("http://mtfocus.io", "MT Focus Technology")
        info.Developers = ["Likit Preeyanon\nEmail: likit.pre@mahidol.edu"]
        info.License = wordwrap("MIT open source license",
                                500, wx.ClientDC(self.preview_panel))
        wx.adv.AboutBox(info)

    def onConnectDbMenuItemClick(self, event):
        if not self.dbengine:
            with wx.FileDialog(None, "Open data file",
                               wildcard='SQLite files (*.sqlite;*.db)|*.sqlite;*.db',
                               style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) \
                    as fileDialog:
                if fileDialog.ShowModal() == wx.ID_CANCEL:
                    return
                else:
                    self.db_filepath = fileDialog.GetPath()

        if self.db_filepath:
            self.dbfile_lbl.SetLabelText('Database filepath: {} CONNECTED'.format(self.db_filepath))
            self.dbengine = sa.create_engine('sqlite:///{}'.format(self.db_filepath))
            try:
                rf = pandas.read_sql_table('records', con=self.dbengine)
                df = pandas.read_sql_table('drugs', con=self.dbengine)
                df_pivot = df.pivot(index='sur_key', columns='drug', values='sensitivity')
                joined = rf.join(df_pivot)
            except ValueError:
                return wx.MessageBox(caption='Database Error',
                            message='Database schema not valid. The "Data" table not available.')

            self.datafile_lbl.SetLabelText("Data filepath: {}".format(self.data_filepath))
            self.data_loaded = True
            self.data_grid_box_sizer.Remove(0)
            self.data_grid.Destroy()
            self.data_grid = DataGrid(self.preview_panel)
            col_ = list(joined.columns)
            col_.remove('sur_key')
            # temporarily set the table
            self.field_attr.update_from_dataframe(joined[col_])

            metadata = pandas.read_sql_table('metadata', con=self.dbengine)
            self.profile_filepath = metadata.tail(1)['profile'].tolist()[0]
            updated_joined = self.load_profile_from_filepath(joined[col_])
            self.profile_lbl.SetLabelText("Profile filepath: {}".format(self.profile_filepath))

            self.field_attr_list.ClearAll()
            self.refresh_field_attr_list_column()
            self.update_edit_panel(self.field_attr.iget_column(0))

            self.data_grid.set_table(updated_joined)
            self.data_grid.AutoSizeColumns()
            self.data_grid_box_sizer.Add(self.data_grid, 1, flag=wx.EXPAND | wx.ALL)
            self.data_grid_box_sizer.Layout()  # repaint the sizer

            # onFieldAttrListItemSelected() requires columns from data_grid.table.df
            # TODO: might need to refactor this code
            self.field_attr_list.Select(0)

            self.saveProfileItem.Enable(True)
            self.loadProfileItem.Enable(True)
            self.organismItem.Enable(True)
            self.exportToExcelMenuItem.Enable(True)
            self.saveToFlatDbMenuItem.Enable(True)
            self.addToFlatDbMenuItem.Enable(True)

            self.saveToDatabaseMenuItem.Enable(True)
            self.appendToDatabaseMenuItem.Enable(True)
            self.createFieldItem.Enable(True)

    def onDisconnectDbMenuItemClick(self, event):
        if self.dbengine:
            self.dbengine = None
            self.dbfile_lbl.SetLabelText('Database filepath: NOT CONNECTED')
            self.profile_lbl.SetLabelText('Profile filepath: None')
            #TODO: find out the better way to reset the data grid
            self.data_grid_box_sizer.Remove(0)
            self.data_grid.Destroy()
            self.data_grid = DataGrid(self.preview_panel)
            self.data_grid.set_table(pandas.DataFrame())
            self.data_grid.AutoSizeColumns()
            self.data_grid_box_sizer.Add(self.data_grid, 1, flag=wx.EXPAND | wx.ALL)
            self.data_grid_box_sizer.Layout()  # repaint the sizer
            self.saveToFlatDbMenuItem.Enable(False)
            self.addToFlatDbMenuItem.Enable(False)

    def onBiogramDbMenuItemClick(self, event):
        dwengine = None
        dwmeta = sa.MetaData()
        with wx.FileDialog(None, "Choose a flat SQLite data file",
                           wildcard='SQLite files (*.sqlite;*.db)|*.sqlite;*.db',
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) \
                as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            else:
                dw_filepath = fileDialog.GetPath()
        if dw_filepath:
            dwengine = sa.create_engine('sqlite:///{}'.format(dw_filepath))

        if dwengine:
            try:
                dwconn = dwengine.connect()
                metadata = pandas.read_sql_table('metadata', con=dwengine)
            except ValueError:
                with wx.MessageDialog(self, message='Please choose another database file.',
                                      caption='Database failed to connect.') as md:
                    md.ShowModal()
                    return
            else:
                profile_filepath = metadata.tail(1)['profile'].tolist()[0]

            try:
                profile = json.loads(open(profile_filepath, 'r').read())
            except IOError:
                # Try finding a profile file from the data directory
                try:
                    _datapath = os.path.dirname(dw_filepath)
                    _profile_filename = os.path.split(profile_filepath)[-1]
                    profile = json.loads(open(os.path.join(_datapath,
                                                           _profile_filename), 'r').read())
                except IOError:
                    with wx.MessageDialog(self,
                                          message='Cannot read from {}. It does not exist.'\
                                          .format(profile_filepath),
                                          caption='Profile not found') as md:
                        md.ShowModal()
                    return

            date_column = None
            for column in profile['data']:
                if profile['data'][column]['date'] and \
                        profile['data'][column]['keep']:
                    date_column = profile['data'][column]['alias']
                    break

            try:
                fact_table = sa.Table('facts', dwmeta, autoload=True, autoload_with=dwengine)
            except ValueError:
                return wx.MessageBox(message=('Cannot retrieve data from {}.'
                                               '\nThe database must contain the fact table.'.format(dw_filepath)),
                                               caption='Database is not valid.')

            fact_columns = fact_table.c.keys()
            fact_columns.remove('added_at')
            if ('sensitivity' not in fact_columns) or ('drug' not in fact_columns) \
                    or ('drugGroup' not in fact_columns):
                return wx.MessageBox(message='Please choose another database file.',
                                        caption='Database schema is not valid.')

            included_fields = list(fact_columns)
            included_fields.remove('sensitivity')
            included_fields.remove('drug')
            included_fields.remove('drugGroup')

            dlg = IndexFieldList(choices=included_fields)

            info = {}
            info['profile filepath'] = [profile_filepath]
            info['data source'] = [dw_filepath]

            if dlg.ShowModal() == wx.ID_OK:
                rawDataIncluded = dlg.rawDataIncluded.IsChecked()
                if dlg.chlbox.CheckedItems:
                    indexes = [included_fields[i] for i in dlg.indexes]
                    query_columns = [fact_table.c[idx] for idx in indexes] + [
                                                fact_table.c.drugGroup,
                                                fact_table.c.drug,
                                                fact_table.c.sensitivity, 
                                                sa.func.count(fact_table.c.sensitivity)
                                            ]
                    s = sa.select(query_columns).where(fact_table.c.sensitivity!='-')
                    source_data = pandas.read_sql_table('facts', con=dwconn)
                    if date_column:
                        if not dlg.all.IsChecked():
                            startdate = map(int, dlg.startDatePicker.GetValue().FormatISODate().split('-'))
                            enddate = map(int, dlg.endDatePicker.GetValue().FormatISODate().split('-'))
                            startdate = pandas.Timestamp(*startdate)
                            enddate = pandas.Timestamp(*enddate)
                            #df_filter = df[(df[date_column] >= startdate) & (df[date_column] <= enddate)]
                            s = s.where(sa.and_(fact_table.c[date_column]>=startdate, fact_table.c[date_column]<=enddate))
                            info['startdate'] = [startdate]
                            info['enddate'] = [enddate]
                            source_data = source_data[(source_data[date_column]>=startdate)
                                                      & (source_data[date_column]<=enddate)]

                    for index in indexes:
                        s = s.group_by(fact_table.c[index])

                    s = s.group_by(fact_table.c.drug)
                    s = s.group_by(fact_table.c.sensitivity)

                    ncutoff = dlg.ncutoff.GetValue()
                    thread = Thread(target=self.generate_antibiogram, args=(s, dwengine, indexes, ncutoff))
                    thread.start()
                    result = NotificationBox(self, caption='Generate Antibiogram',
                                         message='Calculating antibiogram, please wait...').ShowModal()

                    if result == 1:
                        return wx.MessageBox(caption='Empty Antibiogram',
                                                message=('The antibiogram contains no data.\n'
                                                            'Please adjust the minimum number of isolates.'))
                    elif result > 1:
                        return wx.MessageBox(caption='Unknown Error Occurred',
                                                message=('Program failed to calculate the antibiogram'
                                                            'due to data integrity problem.'))

                    with wx.FileDialog(None, "Specify the output file",
                                       wildcard='Excel files (*.xlsx)|*.xlsx',
                                       style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) \
                            as fileDialog:
                        if fileDialog.ShowModal() != wx.ID_CANCEL:
                            excel_filepath = fileDialog.GetPath()
                            writer = pandas.ExcelWriter(excel_filepath)
                            self.biogram_data['biogram_total'].fillna(0).to_excel(writer, 'total')
                            self.biogram_data['biogram_s'].to_excel(writer, 'count_s')
                            self.biogram_data['biogram_ri'].to_excel(writer, 'count_ir')
                            self.biogram_data['biogram_s_pct'].to_excel(writer, 'percent_s')
                            self.biogram_data['biogram_ri_pct'].to_excel(writer, 'percent_ir')
                            self.biogram_data['biogram_narst_s'].to_excel(writer, 'narst_s')
                            self.biogram_data['biogram_narst_r'].to_excel(writer, 'narst_ir')

                            if rawDataIncluded:
                                source_data.to_excel(writer, 'source', index=False)

                            pandas.DataFrame(info).to_excel(writer, 'info', index=False)
                            writer.save()

                            with wx.MessageDialog(None, message='Antibiogram is generated.', caption='Finished',
                                                  style=wx.OK | wx.CENTER) as msgDialog:
                                msgDialog.ShowModal()
                else:
                    with wx.MessageDialog(None,
                                          message='Please choose at least one column as an index of the antibiogram.',
                                          caption='No indexes specified.',
                                          style=wx.OK | wx.CENTER) as msgDialog:
                        msgDialog.ShowModal()


    def generate_antibiogram(self, command, engine, indexes, ncutoff=0, heatmap=False):
        # TODO: move biogram_data into the init method
        self.biogram_data = {}

        def check_cutoff(x, cutoff=ncutoff):
            if x < cutoff:
                return None
            else:
                return x

        connection = engine.connect()
        rp = connection.execute(command)
        if not heatmap:
            columns = indexes + ['drug_group', 'drug', 'result', 'count']
        else:
            columns = indexes + ['drug', 'result', 'count']
        df = pandas.DataFrame(rp.fetchall(), columns=columns)

        if len(df) > 0:
            if not heatmap:
                total = df.pivot_table(index=indexes,
                                       columns=['drug_group', 'drug'],
                                       aggfunc='sum')
            else:
                total = df.pivot_table(index=indexes,
                                       columns=['drug'], aggfunc='sum')

            flt_total = total.applymap(check_cutoff)
            sens = df[df['result'] == 'S']
            if not heatmap:
                sens = sens.pivot_table(index=indexes, columns=['drug_group','drug'])
            else:
                sens = sens.pivot_table(index=indexes, columns=['drug']).fillna(0)
            resists = total - sens

            self.biogram_data['biogram_total'] = total
            self.biogram_data['biogram_s'] = sens
            self.biogram_data['biogram_ri'] = resists
            sens_pct = round((sens / flt_total) * 100, 2)
            resists_pct =  round((resists / flt_total) * 100, 2)
            self.biogram_data['biogram_s_pct'] = sens_pct.fillna('')
            self.biogram_data['biogram_ri_pct'] = resists_pct.fillna('')
            biogram_narst_s = round(sens_pct).fillna('').applymap(str) + \
                              " (" + flt_total.applymap(lambda x: '' if pandas.isna(x) else '{:.0f}'.format(x)) + ")"
            biogram_narst_r = round(resists_pct).fillna('').applymap(str) + \
                              " (" + flt_total.applymap(lambda x: '' if pandas.isna(x) else '{:.0f}'.format(x)) + ")"
            biogram_narst_s = biogram_narst_s.applymap(lambda x: '' if x == ' ()' else x)
            biogram_narst_r = biogram_narst_r.applymap(lambda x: '' if x == ' ()' else x)
            self.biogram_data['biogram_narst_s'] = biogram_narst_s
            self.biogram_data['biogram_narst_r'] = biogram_narst_r
            wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=0)
        else:
            wx.CallAfter(dispatcher.send, CLOSE_DIALOG_SIGNAL, rc=1)

    def onBiogramHeatmapMenuItemClick(self, event):
        dwengine = None
        dwmeta = sa.MetaData()
        with wx.FileDialog(None, "Choose a flat SQLite data file",
                           wildcard='SQLite files (*.sqlite;*.db)|*.sqlite;*.db',
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) \
                as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            else:
                dw_filepath = fileDialog.GetPath()
        if dw_filepath:
            dwengine = sa.create_engine('sqlite:///{}'.format(dw_filepath))

        if dwengine:
            try:
                dwconn = dwengine.connect()
                metadata = pandas.read_sql_table('metadata', con=dwengine)
            except ValueError:
                with wx.MessageDialog(self, message='Please choose another database file.',
                                      caption='Database failed to connect.') as md:
                    md.ShowModal()
                    return
            else:
                profile_filepath = metadata.tail(1)['profile'].tolist()[0]

            try:
                profile = json.loads(open(profile_filepath, 'r').read())
            except IOError:
                # Try finding a profile file from the data directory
                try:
                    _datapath = os.path.dirname(dw_filepath)
                    _profile_filename = os.path.split(profile_filepath)[-1]
                    profile = json.loads(open(os.path.join(_datapath,
                                                           _profile_filename), 'r').read())
                except IOError:
                    with wx.MessageDialog(self, message='Cannot read from {}. It does not exist.'.format(profile_filepath),
                                      caption='Profile not found') as md:
                        md.ShowModal()
                    return

            date_column = None
            for column in profile['data']:
                if profile['data'][column]['date'] and \
                        profile['data'][column]['keep']:
                    date_column = profile['data'][column]['alias']
                    break

            try:
                fact_table = sa.Table('facts', dwmeta, autoload=True, autoload_with=dwengine)
            except ValueError:
                return wx.MessageBox(message=('Cannot retrieve data from {}.'
                                              '\nThe database must contain the fact table.'.format(dw_filepath)),
                                     caption='Database is not valid.')

            fact_columns = fact_table.c.keys()
            fact_columns.remove('added_at')
            if ('sensitivity' not in fact_columns) or ('drug' not in fact_columns) \
                    or ('drugGroup' not in fact_columns):
                return wx.MessageBox(message='Please choose another database file.',
                                     caption='Database schema is not valid.')

            included_fields = list(fact_columns)
            included_fields.remove('sensitivity')
            included_fields.remove('drug')
            included_fields.remove('drugGroup')

            dlg = HeatmapFieldList(choices=included_fields)

            info = {}
            info['profile filepath'] = [profile_filepath]
            info['data source'] = [dw_filepath]

            if dlg.ShowModal() == wx.ID_OK:
                if dlg.chlbox.CheckedItems:
                    if len(dlg.indexes) > 1:
                        with wx.MessageDialog(None,
                                              "Only single field is supported"
                                              " for this version.",
                                              "Multiple fields were selected.",
                                               wx.OK) as msg:
                            if msg.ShowModal() == wx.ID_OK:
                                return

                    indexes = ['organism_name'] + [included_fields[i] for i in dlg.indexes]
                    query_columns = [fact_table.c[idx] for idx in indexes] + [
                        fact_table.c.drug,
                        fact_table.c.sensitivity,
                        sa.func.count(fact_table.c.sensitivity)
                    ]
                    s = sa.select(query_columns)
                    source_data = pandas.read_sql_table('facts', con=dwconn)
                    if date_column:
                        if not dlg.all.IsChecked():
                            startdate = map(int, dlg.startDatePicker.GetValue().FormatISODate().split('-'))
                            enddate = map(int, dlg.endDatePicker.GetValue().FormatISODate().split('-'))
                            startdate = pandas.Timestamp(*startdate)
                            enddate = pandas.Timestamp(*enddate)
                            #df_filter = df[(df[date_column] >= startdate) & (df[date_column] <= enddate)]
                            s = s.where(sa.and_(fact_table.c[date_column]>=startdate, fact_table.c[date_column]<=enddate))
                            info['startdate'] = [startdate]
                            info['enddate'] = [enddate]
                            source_data = source_data[(source_data[date_column]>=startdate)
                                                      & (source_data[date_column]<=enddate)]
                    ncutoff = dlg.ncutoff.GetValue()
                    organisms = sorted(source_data['organism_name'].unique())
                    with wx.SingleChoiceDialog(None, "Select an organism", "Organisms", organisms) as org_dlg:
                        if org_dlg.ShowModal() == wx.ID_OK:
                            select_organism =  org_dlg.GetStringSelection()
                        else:
                            return

                    for index in indexes:
                        s = s.group_by(fact_table.c[index])

                    s = s.group_by(fact_table.c.drug)
                    s = s.group_by(fact_table.c.sensitivity)
                    s = s.where(fact_table.c.organism_name==select_organism)

                    thread = Thread(target=self.generate_antibiogram, args=(s, dwengine, indexes, ncutoff, True))
                    thread.start()
                    result = NotificationBox(self, caption='Generate Antibiogram',
                                             message='Calculating antibiogram, please wait...').ShowModal()

                    if result == 1:
                        return wx.MessageBox(caption='Empty Antibiogram',
                                             message=('The antibiogram contains no data.\n'
                                                      'Please adjust the minimum number of isolates.'))
                    elif result > 1:
                        return wx.MessageBox(caption='Unknown Error Occurred',
                                             message=('Program failed to calculate the antibiogram'
                                                      'due to data integrity problem.'))
                    heatmap_df = self.biogram_data['biogram_s_pct']['count'].droplevel(level=0)
                    self.plot_heatmap(heatmap_df, select_organism)
                    return

    def plot_heatmap(self, df, title):
        import numpy
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('WXAgg')
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
        import seaborn as sns

        plt.rcParams['font.family'] = 'TH Sarabun New'

        df = df.replace(r'^\s*$', numpy.nan, regex=True)
        df.dropna(1, how='all', inplace=True)
        df = df.fillna(120)
        if df.empty:
            return wx.MessageBox(caption='Distance matrix is empty.',
                                 message=('The plot could not be created because the data table is empty.'))

        wildcard = "PNG (*.png)|*.png"
        filepath = self.data_filepath or os.getcwd()
        with wx.FileDialog(None, "Choose a file to save a plot.", filepath,
                           "", wildcard, wx.FC_SAVE) as file_dlg:
            if file_dlg.ShowModal() == wx.ID_CANCEL:
                return
            else:
                filepath = file_dlg.GetPath()

        try:
            sns.clustermap(df, cmap=sns.diverging_palette(20, 220, n=7))
            plt.suptitle(title)
            plt.savefig(filepath)
        except:
            return wx.MessageBox(caption='Plotting Failed.',
                                 message=('The plot could not be generated or saved to the file.'))
        else:
            return wx.MessageBox(caption='Finished.',
                             message=('The plot has been saved to the file.'))
